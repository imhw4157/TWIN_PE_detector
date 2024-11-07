import sys
import os
import regex
import pandas as pd
import openpyxl as op
import numpy as np

from collections import Counter
import multiprocessing

from Bio.SeqIO.QualityIO import FastqGeneralIterator
from Bio.Seq import Seq
from Bio.SeqIO import write
from Bio.Seq import reverse_complement
from Bio import Align

from Bio import pairwise2
from Bio.pairwise2 import format_alignment


from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment





ind = os.getcwd()

f_list = open('target_list.txt').readlines()

first_line_list = f_list[0][:-1].split('\t')

minumum_num=int(first_line_list[1])
comparison_range=int(first_line_list[3])
indicator_range=int(first_line_list[5])
window_range=int(first_line_list[7])







# minimum_frequency : Time Complexity O(1)

def minimum_frequency(records, n=2):
    seq_form_fastq = [record[1] for record in records]  
    counts = Counter(seq_form_fastq)    
    minimum_records = [record for record in records if counts[record[1]] >= n] 
    return minimum_records 



# indicator selection : w/ mismatch

def pick_indicator_with_mismatch(seq_input, left_indicator, right_indicator): # seq_input type = minimum_records
    indicator_records=[]
    if spacer_F_seq in wt_ref_seq:
        for seq in seq_input:
            is_WT_L = regex.search(rf"({left_indicator}){{s<={2}}}", seq[1]) # regex fuzzy matching LI
            is_WT_R = regex.search(rf"({right_indicator}){{s<={2}}}", seq[1]) # regex fuzzy matching RI
            if is_WT_L is not None and is_WT_R is not None:
                nomean=['A','B','C']
                nomean[1]=seq[1][is_WT_L.start():is_WT_R.end()]
                indicator_records.append(nomean)
                
    elif spacer_F_seq in wt_ref_seq_complement:
        for seq in seq_input:
            is_WT_L = regex.search(rf"({left_indicator}){{s<={2}}}", str(Seq(seq[1]).reverse_complement())) # regex fuzzy matching LI
            is_WT_R = regex.search(rf"({right_indicator}){{s<={2}}}", str(Seq(seq[1]).reverse_complement())) # regex fuzzy matching RI
            if is_WT_L is not None and is_WT_R is not None:
                nomean=['A','B','C']
                nomean[1]=str(Seq(seq[1]).reverse_complement())[is_WT_L.start():is_WT_R.end()]
                indicator_records.append(nomean)

    return indicator_records 






# count mutagen

def count_mutagen(seq_input, ref_seq, inerstion, WT_marker_F, WT_marker_R): # seq_input type = indicator_records
    cnt = {'dimer' : 0, 'ins' : 0, 'del' : 0, 'sub' : 0, 'WT' :0, 'edit': 0, 'unsorted' : 0, 'indel edit' : 0}
    
    for seq in seq_input: 
        if len(seq[1]) < len(ref_seq)*0.15: # 0.3 is a threshold caution : it can be changed
            cnt['dimer'] += 1

        else:
            if WT_marker_F in seq[1] and WT_marker_R in seq[1]:
                
                if len(seq[1]) > len(ref_seq):
                    cnt['ins'] += 1

                elif len(seq[1]) < len(ref_seq):
                    cnt['del'] += 1

                else:
                    cnt['WT'] += 1
                
            elif WT_marker_F_half + insertion + WT_marker_R_half in seq[1]:
                cnt['edit'] += 1
               
            
            elif insertion in seq[1]:
                cnt['indel edit'] += 1
                
            elif len(seq[1]) == len(ref_seq):
                cnt['sub'] += 1
                
            elif len(seq[1]) > len(ref_seq):
                cnt['ins'] += 1
            
            elif len(seq[1]) < len(ref_seq):
                cnt['del'] += 1
                
            else:
                cnt['unsorted'] += 1
                
    print(cnt)
    return cnt
    
    
   








###  ###








wb = Workbook()
sheet = wb.active
sheet.title = "Analysis Result"

thisis2=2
thisis1=1
thisis0=0

for t in f_list[1:]:    
    each_line_list = t.split('\t')

    if len(each_line_list) <= 1:
        continue









##################################  ##################################

    

    wt_ref = each_line_list[0].strip()
    spacer_F = each_line_list[1].strip()
    spacer_R = each_line_list[2].strip()
    insertion = each_line_list[3].strip()
    direc = each_line_list[4].strip()
    inedx = each_line_list[5].strip()
    name = each_line_list[6].strip()


    wt_ref_seq = wt_ref.upper()
    spacer_F_seq = spacer_F.upper()
    spacer_R_seq = spacer_R.upper()
    spacer_R_seq_complement=str(Seq(spacer_R_seq).reverse_complement())
    insertion_seq = insertion.upper()
    wt_ref_seq_complement=str(Seq(wt_ref_seq).reverse_complement())






    if spacer_F_seq in wt_ref_seq:
        position_nick_F = wt_ref_seq.find(spacer_F_seq) + 17
        position_nick_R = wt_ref_seq.find(spacer_R_seq_complement) + 3

        WT_marker_F = wt_ref_seq[position_nick_F-5:position_nick_F+5]
        WT_marker_R = wt_ref_seq[position_nick_R-5:position_nick_R+5]
        
        WT_marker_F_half = wt_ref_seq[position_nick_F-5:position_nick_F]
        WT_marker_R_half = wt_ref_seq[position_nick_R:position_nick_R+5]

        if position_nick_F-comparison_range >= 0:
            left_indicator = wt_ref_seq[position_nick_F-comparison_range : position_nick_F-comparison_range+indicator_range]
        else:
            left_indicator = wt_ref_seq[:indicator_range]

        if position_nick_R+comparison_range < len(wt_ref_seq) :
            right_indicator = wt_ref_seq[position_nick_R+comparison_range-indicator_range : position_nick_R+comparison_range]
        else:
            right_indicator = wt_ref_seq[-indicator_range:]
        
        left_index = wt_ref_seq.find(left_indicator)
        right_index = wt_ref_seq.find(right_indicator)
        wt_ref_seq = wt_ref_seq[left_index:right_index+len(right_indicator)]
        position_nick_F = wt_ref_seq.find(spacer_F_seq) + 17
        position_nick_R = wt_ref_seq.find(spacer_R_seq_complement) + 3



        
    elif spacer_F_seq in wt_ref_seq_complement:
        position_nick_F = wt_ref_seq_complement.find(spacer_F_seq) + 17
        position_nick_R = wt_ref_seq_complement.find(spacer_R_seq_complement) + 3

        WT_marker_F = wt_ref_seq_complement[position_nick_F-5:position_nick_F+5]
        WT_marker_R = wt_ref_seq_complement[position_nick_R-5:position_nick_R+5]
        
        WT_marker_F_half = wt_ref_seq_complement[position_nick_F-5:position_nick_F]
        WT_marker_R_half = wt_ref_seq_complement[position_nick_R:position_nick_R+5]

        if position_nick_F-comparison_range >= 0:
            left_indicator = wt_ref_seq_complement[position_nick_F-comparison_range : position_nick_F-comparison_range+indicator_range]
        else:
            left_indicator = wt_ref_seq_complement[:indicator_range]

        if position_nick_R+comparison_range < len(wt_ref_seq) :
            right_indicator = wt_ref_seq_complement[position_nick_R+comparison_range-indicator_range : position_nick_R+comparison_range]
        else:
            right_indicator = wt_ref_seq_complement[-indicator_range:]
            
            
        left_index = wt_ref_seq_complement.find(left_indicator)
        right_index = wt_ref_seq_complement.find(right_indicator)
        wt_ref_seq_complement= wt_ref_seq_complement[left_index:right_index+len(right_indicator)]
        position_nick_F = wt_ref_seq_complement.find(spacer_F_seq) + 17
        position_nick_R = wt_ref_seq_complement.find(spacer_R_seq_complement) + 3

        




    else:
        print("No spacer found.")
        print(f"\u25B6 \u25B6 # {inedx} analysis failed.")
        print("--------------------------------------------------------------------------")

        continue





##################################  ##################################


        



##################################  ##################################


    file_name = str(inedx) + '.fastqjoin'
    os.chdir(ind + '/' + direc)



################################## fastqjoin 파일 탐색 후 open ##################################









    records = [] 
    with open(file_name, "r") as handle:
        for record in FastqGeneralIterator(handle): 
            records.append(record)



    os.chdir(ind)



    record_minimum=minimum_frequency(records, n=minumum_num)
    record_indicator=pick_indicator_with_mismatch(record_minimum, left_indicator, right_indicator)
    print(f"#{inedx} : indicator matched = {len(record_indicator)}")

    if len(record_indicator) == 0:
        print(f"\u25B6 # {inedx} has no read.")
        print("--------------------------------------------------------------------------")
        continue
    
    
    
    
    
    

    else:
        if spacer_F_seq in wt_ref_seq:
            record_mutagen=count_mutagen(record_indicator, wt_ref_seq, insertion, WT_marker_F, WT_marker_R)

            
        
        else:
            record_mutagen=count_mutagen(record_indicator, wt_ref_seq, insertion, WT_marker_F, WT_marker_R)

                        
        
                        


        




    
                        
        cnt=record_mutagen
        
        Total_count = cnt['ins']+cnt['del']+cnt['WT']+cnt['sub']+cnt['edit']+cnt['unsorted']+cnt['indel edit']




        header0 = {
            'Memo_name': ['Memo'],
            'index_name' : ['index'],
            'Min.Freq_name': ['Min.Freq'],
            'wo_name' : ['Total (w/o dimer)'],
            'WT_name' : ['WT'],
            'ins' : ['ins'],
            'del' : ['del'],
            'sub' : ['sub'],
            'edit' : ['edit'],
            'indel edit' : ['indel edit'],
            'unsorted' : ['unsorted']
        }




        header = {
            'file name': [str(name)],
            'index' : [str(inedx)],
            'Min.Freq.': [str(len(record_indicator))],
            'wo': [str(Total_count)],
            'WT': [str(cnt['WT'])],
            'ins': [str(cnt['ins'])],
            'del': [str(cnt['del'])],
            'sub': [str(cnt['sub'])],
            'edit': [str(cnt['edit'])],
            'indel edit': [str(cnt['indel edit'])],
            'unsorted': [str(cnt['unsorted'])]
        }


        df_header=pd.DataFrame(header)
        df_header0=pd.DataFrame(header0)





        # ?  ? ? 일??? 기
        
        
        for r_idx, row in enumerate(dataframe_to_rows(df_header0, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        for r_idx, row in enumerate(df_header.itertuples(index=False), start=1):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=thisis2 + r_idx, column=c_idx, value=value)



        thisis2 += 2
        thisis1 += 2
        thisis0 += 2
        
        print(f"\u25B6 #{file_name} analysis completed.")
        print("--------------------------------------------------------------------------")

    # exprot
    wb.save('sub.xlsx')

print('Jobs done! made by HW')
